# -*- coding: utf-8 -*-
"""
@Time    : 2023/2/15 17:55
@Author  : itlubber
@Site    : itlubber.art
"""
__version__ = "0.1.39"

__all__ = (
    "__version__"
    , "FeatureSelection", "FeatureImportanceSelector", "StepwiseSelection", "Combiner", "WOETransformer"
    , "ITLubberLogisticRegression", "ScoreCard", "Rule", "DecisionTreeRuleExtractor", "ruleset_report", "sawpin_badrate_prediction_by_score", "bin_table_badrate_prediction", "swapin_report", "swapout_report"
    , "Pipeline", "KS", "AUC", "PSI", "F1", "FeatureUnion", "make_pipeline", "make_union"
    , "init_logger", "init_setting", "init_font_for_excel", "get_excel_font_name", "load_pickle", "save_pickle", "germancredit"
    , "ColorScaleRule", "get_column_letter", "column_index_from_string", "seed_everything"
    , "feature_bins", "feature_bin_stats", "feature_efficiency_analysis", "extract_feature_bin", "inverse_feature_bins", "sample_lift_transformer", "feature_describe", "groupby_feature_describe", "feature_summary"
    , "bin_plot", "corr_plot", "ks_plot", "hist_plot", "psi_plot", "csi_plot", "dataframe_plot", "distribution_plot"
    , "bin_trend_plot", "batch_bin_trend_plot", "bin_overdues_plot"
    , "ExcelWriter", "dataframe2excel", "auto_eda_sweetviz", "auto_data_testing_report", "QuickModelReport", "auto_model_report"
    , "RFE", "RFECV", "SelectKBest", "SelectFromModel", "GenericUnivariateSelect", "NumExprDerive"
    , "StandardScoreTransformer", "NPRoundStandardScoreTransformer", "RoundStandardScoreTransformer", "BoxCoxScoreTransformer"
    , "TypeSelector", "RegexSelector", "ModeSelector", "NullSelector", "InformationValueSelector", "LiftSelector"
    , "VarianceSelector", "VIFSelector", "CorrSelector", "PSISelector", "NullImportanceSelector", "TargetPermutationSelector", "ExhaustiveSelector"
)


def __getattr__(name):
    """Lazy imports for better import performance."""
    from sklearn.pipeline import Pipeline, FeatureUnion, make_pipeline, make_union
    if name in ("Pipeline", "FeatureUnion", "make_pipeline", "make_union"):
        return locals()[name]
    
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter, column_index_from_string
    if name in ("ColorScaleRule", "get_column_letter", "column_index_from_string"):
        return locals()[name]
    
    from toad.metrics import KS, AUC, F1, PSI
    if name in ("KS", "AUC", "F1", "PSI"):
        return locals()[name]
    
    from .logger import init_logger
    if name == "init_logger":
        return init_logger
    
    from .utils import init_setting, init_font_for_excel, get_excel_font_name, load_pickle, save_pickle, germancredit, seed_everything, feature_bins, extract_feature_bin, inverse_feature_bins, sample_lift_transformer, feature_describe, groupby_feature_describe, feature_summary
    if name in ("init_setting", "init_font_for_excel", "get_excel_font_name", "load_pickle", "save_pickle", "germancredit", "seed_everything", "feature_bins", "extract_feature_bin", "inverse_feature_bins", "sample_lift_transformer", "feature_describe", "groupby_feature_describe", "feature_summary"):
        return locals()[name]
    
    from .processing import FeatureSelection, FeatureImportanceSelector, StepwiseSelection, Combiner, WOETransformer, feature_bin_stats, feature_efficiency_analysis
    if name in ("FeatureSelection", "FeatureImportanceSelector", "StepwiseSelection", "Combiner", "WOETransformer", "feature_bin_stats", "feature_efficiency_analysis"):
        return locals()[name]
    
    from .model import ITLubberLogisticRegression, ScoreCard
    if name in ("ITLubberLogisticRegression", "ScoreCard"):
        return locals()[name]
    
    from .excel_writer import ExcelWriter, dataframe2excel
    if name in ("ExcelWriter", "dataframe2excel"):
        return locals()[name]
    
    from .auto_eda import auto_eda_sweetviz
    if name == "auto_eda_sweetviz":
        return auto_eda_sweetviz
    
    from .auto_report import auto_data_testing_report
    if name == "auto_data_testing_report":
        return auto_data_testing_report
    
    from .model_report import QuickModelReport, auto_model_report
    if name in ("QuickModelReport", "auto_model_report"):
        return locals()[name]
    
    from .rule import Rule, ruleset_report, sawpin_badrate_prediction_by_score, bin_table_badrate_prediction, swapin_report, swapout_report
    if name in ("Rule", "ruleset_report", "sawpin_badrate_prediction_by_score", "bin_table_badrate_prediction", "swapin_report", "swapout_report"):
        return locals()[name]
    
    from .rule_extraction import DecisionTreeRuleExtractor
    if name == "DecisionTreeRuleExtractor":
        return DecisionTreeRuleExtractor
    
    from .feature_engineering import NumExprDerive
    if name == "NumExprDerive":
        return NumExprDerive
    
    from .feature_selection import RFE, RFECV, SelectKBest, SelectFromModel, GenericUnivariateSelect, TypeSelector, RegexSelector, ModeSelector, NullSelector, InformationValueSelector, LiftSelector, VarianceSelector, VIFSelector, CorrSelector, PSISelector, NullImportanceSelector, TargetPermutationSelector, ExhaustiveSelector
    if name in ("RFE", "RFECV", "SelectKBest", "SelectFromModel", "GenericUnivariateSelect", "TypeSelector", "RegexSelector", "ModeSelector", "NullSelector", "InformationValueSelector", "LiftSelector", "VarianceSelector", "VIFSelector", "CorrSelector", "PSISelector", "NullImportanceSelector", "TargetPermutationSelector", "ExhaustiveSelector"):
        return locals()[name]
    
    from .scorecard import StandardScoreTransformer, NPRoundStandardScoreTransformer, RoundStandardScoreTransformer, BoxCoxScoreTransformer
    if name in ("StandardScoreTransformer", "NPRoundStandardScoreTransformer", "RoundStandardScoreTransformer", "BoxCoxScoreTransformer"):
        return locals()[name]
    
    from .utils import bin_plot, corr_plot, ks_plot, hist_plot, psi_plot, csi_plot, dataframe_plot, distribution_plot, bin_trend_plot, batch_bin_trend_plot, bin_overdues_plot
    if name in ("bin_plot", "corr_plot", "ks_plot", "hist_plot", "psi_plot", "csi_plot", "dataframe_plot", "distribution_plot", "bin_trend_plot", "batch_bin_trend_plot", "bin_overdues_plot"):
        return locals()[name]
    
    raise AttributeError(f"module {__name__!r} has no attribute {name!r}")